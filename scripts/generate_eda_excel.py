"""
EDA Summary Excel Generator
Generates a comprehensive EDA Excel workbook for transaction_with_graph_features.parquet
Sheets: Overview, Data Dictionary, Feature Formulas, Descriptive Stats,
        Missing Values, Graph Features EDA, Categorical Distributions,
        Rule Features, Temporal Analysis, Fraud Type Analysis, Top Correlations
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
PARQUET_PATH = BASE_DIR / "outputs" / "transaction_with_graph_features.parquet"
OUTPUT_PATH  = BASE_DIR / "outputs" / "EDA_transaction_with_graph_features.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"   # header background
C_MED_BLUE    = "2E75B6"   # sub-header
C_LIGHT_BLUE  = "D6E4F0"   # alternating row
C_ORANGE      = "C55A11"   # alert / critical
C_YELLOW      = "FFD966"   # medium severity
C_GREEN       = "70AD47"   # good / low
C_LIGHT_GREY  = "F2F2F2"   # alternating row 2
C_RED_LIGHT   = "FCE4D6"
C_WHITE       = "FFFFFF"
C_TEAL        = "00B0F0"

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def hdr_font(bold=True, size=11, color=C_WHITE):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def body_font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="B8B8B8")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, row_num, values, col_start=1,
                 bg=C_DARK_BLUE, fg=C_WHITE, size=11):
    for i, val in enumerate(values, start=col_start):
        c = ws.cell(row=row_num, column=i, value=val)
        c.font = hdr_font(size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def apply_subheader(ws, row_num, values, col_start=1, bg=C_MED_BLUE):
    for i, val in enumerate(values, start=col_start):
        c = ws.cell(row=row_num, column=i, value=val)
        c.font = hdr_font(size=10, color=C_WHITE)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def write_df(ws, df, start_row, start_col=1, alt_colors=True):
    """Write a DataFrame to ws starting at start_row, start_col. Returns next free row."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        excel_row = start_row + r_idx
        bg = C_LIGHT_BLUE if (alt_colors and r_idx % 2 == 0) else C_WHITE
        for c_idx, val in enumerate(row, start=start_col):
            c = ws.cell(row=excel_row, column=c_idx, value=val)
            c.font = body_font()
            c.fill = fill(bg)
            c.alignment = left()
            c.border = thin_border()
    return start_row + len(df)

def set_col_widths(ws, widths):
    """widths: list of (col_letter_or_idx, width)"""
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def title_cell(ws, row, col, text, size=14, color=C_DARK_BLUE):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=size, color=color, name="Calibri")
    c.alignment = left()

def section_cell(ws, row, col, text, bg=C_MED_BLUE):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=10, color=C_WHITE, name="Calibri")
    c.fill = fill(bg)
    c.alignment = left()
    c.border = thin_border()

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────
print("Loading parquet …")
df = pd.read_parquet(PARQUET_PATH)
print(f"  Shape: {df.shape}")

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default blank sheet

# =============================================================================
# SHEET 1 – OVERVIEW
# =============================================================================
def sheet_overview(wb, df):
    ws = wb.create_sheet("1_Overview")
    ws.freeze_panes = "B3"

    title_cell(ws, 1, 1, "AML Dataset – EDA Overview", size=16)

    # ── Key metrics block ────────────────────────────────────────────────────
    metrics = [
        ("Total Transactions",        f"{len(df):,}"),
        ("Total Columns / Features",  f"{df.shape[1]}"),
        ("Date Range Start",          str(df["timestamp"].min())[:19]),
        ("Date Range End",            str(df["timestamp"].max())[:19]),
        ("Fraud Transactions (label=1)", f"{df['label'].sum():,}"),
        ("Legitimate (label=0)",      f"{(df['label']==0).sum():,}"),
        ("Overall Fraud Rate",        f"{df['label'].mean()*100:.2f}%"),
        ("Unique Customers",          f"{df['customer_id'].nunique():,}"),
        ("Unique Sender Accounts",    f"{df['sender_account_id'].nunique():,}"),
        ("Unique Devices",            f"{df['device_id'].nunique():,}"),
        ("Unique Transaction Types",  f"{df['transaction_type'].nunique()}"),
        ("Channels",                  ", ".join(df['channel'].unique())),
        ("Columns with Nulls",        f"{(df.isnull().sum()>0).sum()}"),
        ("Total Missing Cells",       f"{df.isnull().sum().sum():,}"),
    ]

    apply_header(ws, 3, ["Metric", "Value"], bg=C_DARK_BLUE)
    for i, (k, v) in enumerate(metrics, start=4):
        ws.cell(row=i, column=1, value=k).font = body_font(bold=True)
        ws.cell(row=i, column=1).fill = fill(C_LIGHT_BLUE if i%2==0 else C_WHITE)
        ws.cell(row=i, column=2, value=v).font = body_font()
        ws.cell(row=i, column=2).fill = fill(C_LIGHT_BLUE if i%2==0 else C_WHITE)
        for col in [1, 2]:
            ws.cell(row=i, column=col).border = thin_border()
            ws.cell(row=i, column=col).alignment = left()

    # ── Fraud type distribution ───────────────────────────────────────────────
    row = 3 + len(metrics) + 3
    title_cell(ws, row, 1, "Fraud Type Distribution", size=12)
    row += 1
    apply_header(ws, row, ["Fraud Type", "Count", "% of All Txns", "Fraud Rate (of type)"], bg=C_MED_BLUE)
    row += 1
    ft = df["fraud_type"].value_counts().reset_index()
    ft.columns = ["fraud_type","count"]
    ft["pct_all"] = (ft["count"] / len(df) * 100).round(2)
    ft["fraud_rate"] = ft["fraud_type"].map(
        df.groupby("fraud_type")["label"].mean().round(4)
    )
    for i, r in ft.iterrows():
        bg = C_RED_LIGHT if r["fraud_rate"] > 0 else C_WHITE
        for ci, val in enumerate([r["fraud_type"], r["count"], f"{r['pct_all']}%", f"{r['fraud_rate']*100:.1f}%"], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    # ── Channel & txn type summary ───────────────────────────────────────────
    row += 2
    title_cell(ws, row, 1, "Channel Distribution", size=12)
    row += 1
    ch = df.groupby("channel")["label"].agg(["count","sum","mean"]).reset_index()
    ch.columns = ["channel","total","fraud_count","fraud_rate"]
    ch["fraud_rate"] = (ch["fraud_rate"]*100).round(2)
    apply_header(ws, row, ["Channel","Total","Fraud Count","Fraud Rate (%)"], bg=C_MED_BLUE)
    row += 1
    for _, r in ch.iterrows():
        for ci, val in enumerate([r["channel"], r["total"], r["fraud_count"], f"{r['fraud_rate']}%"], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = body_font()
            c.fill = fill(C_LIGHT_BLUE if row%2==0 else C_WHITE)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    set_col_widths(ws, [(1, 40), (2, 25), (3, 20), (4, 20)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 1_Overview")


# =============================================================================
# SHEET 2 – DATA DICTIONARY
# =============================================================================
DATA_DICT = {
    # ── Transaction Core ─────────────────────────────────────────────────────
    "timestamp":             ("datetime",  "Transaction Base",  "Exact date-time of the transaction",                                   "N/A"),
    "transaction_id":        ("string",    "Transaction Base",  "Unique transaction identifier (T-prefixed)",                           "N/A"),
    "customer_id":           ("string",    "Transaction Base",  "Unique customer identifier (C-prefixed)",                              "N/A"),
    "sender_account_id":     ("string",    "Transaction Base",  "Account initiating the transaction (A-prefixed)",                      "N/A"),
    "receiver_account_id":   ("string",    "Transaction Base",  "Destination account; NULL for external/beneficiary transfers",         "N/A"),
    "beneficiary_id":        ("string",    "Transaction Base",  "External beneficiary ID; NULL for intra-bank transfers",               "N/A"),
    "device_id":             ("string",    "Transaction Base",  "Device used to initiate transaction (D-prefixed)",                     "N/A"),
    "amount":                ("float",     "Transaction Base",  "Transaction value in local currency",                                  "N/A"),
    "channel":               ("category",  "Transaction Base",  "Channel: mobile, web, atm, branch",                                   "N/A"),
    "debit_credit":          ("category",  "Transaction Base",  "Whether the transaction is a debit or credit to the sender account",   "N/A"),
    "transaction_type":      ("category",  "Transaction Base",  "Payment instrument: IMPS, NEFT, RTGS, UPI, wallet_transfer, etc.",     "N/A"),
    "cash_flag":             ("binary",    "Transaction Base",  "1 = cash involved (deposit or withdrawal)",                           "cash_flag = 1 if transaction_type in {cash_deposit, cash_withdrawal}"),
    "fraud_type":            ("category",  "Labels",            "Fraud typology: normal | mule_ring | layering | smurfing | ATO | identity_fraud", "Synthetic label; assigned during data generation"),
    "label":                 ("binary",    "Labels",            "Ground truth fraud label: 1 = fraud, 0 = legitimate",                 "label = 0 if fraud_type == 'normal' else 1"),

    # ── Customer / Account Profile ───────────────────────────────────────────
    "avg_balance":           ("float",     "Customer Profile",  "Average account balance (simulated)",                                 "N/A"),
    "account_open_days":     ("int",       "Customer Profile",  "Number of days since account was opened",                            "N/A"),
    "kyc_level":             ("category",  "Customer Profile",  "KYC tier of customer: low | medium | high",                          "N/A"),
    "country_risk":          ("category",  "Customer Profile",  "Risk category of the customer's country of residence",               "N/A"),
    "income_bracket":        ("category",  "Customer Profile",  "Income band: low | medium | high",                                   "N/A"),
    "customer_risk_rating":  ("category",  "Customer Profile",  "Internal CDD risk rating: low | medium | high | very_high",         "N/A"),
    "pep_flag":              ("binary",    "Customer Profile",  "1 = Politically Exposed Person",                                     "N/A"),
    "occupation":            ("category",  "Customer Profile",  "Occupation category (salaried, freelancer, student, etc.)",          "N/A"),
    "industry":              ("category",  "Customer Profile",  "Industry sector of the customer",                                    "N/A"),
    "account_type":          ("category",  "Customer Profile",  "Account category: savings | current | retail | corporate | business","N/A"),
    "account_age_days":      ("int",       "Customer Profile",  "Alias for account_open_days (computed at event time)",               "account_age_days = account_open_days + days elapsed since open date"),

    # ── Device ───────────────────────────────────────────────────────────────
    "device_age_days":       ("int",       "Device",            "Number of days since device was first seen",                         "N/A"),
    "rooted_flag":           ("binary",    "Device",            "1 = device is rooted/jailbroken",                                   "N/A"),
    "os_type":               ("category",  "Device",            "Device OS: android | ios | windows | unknown",                      "N/A"),
    "vpn_flag":              ("binary",    "Device",            "1 = VPN or proxy detected on device",                               "N/A"),
    "emulator_flag":         ("binary",    "Device",            "1 = transaction originated from an emulator",                       "N/A"),

    # ── Beneficiary ──────────────────────────────────────────────────────────
    "beneficiary_type":      ("category",  "Beneficiary",       "Type of external beneficiary: individual | merchant | crypto | offshore", "N/A"),
    "beneficiary_country_risk": ("category","Beneficiary",      "Risk rating of the beneficiary's country",                          "N/A"),
    "high_risk_beneficiary": ("binary",    "Beneficiary",       "1 = beneficiary is in high-risk category (crypto or offshore)",     "high_risk_beneficiary = 1 if beneficiary_type in {crypto, offshore}"),

    # ── Temporal Derived ─────────────────────────────────────────────────────
    "hour":                  ("int",       "Temporal",          "Hour of day (0–23)",                                                 "hour = timestamp.hour"),
    "day_of_week":           ("int",       "Temporal",          "Day of week (0=Mon, 6=Sun)",                                        "day_of_week = timestamp.dayofweek"),
    "date":                  ("string",    "Temporal",          "Calendar date (YYYY-MM-DD)",                                        "date = timestamp.date()"),
    "month":                 ("int",       "Temporal",          "Month (1–12)",                                                      "month = timestamp.month"),
    "week_of_year":          ("int",       "Temporal",          "ISO week of year",                                                  "week_of_year = timestamp.isocalendar().week"),
    "is_night":              ("binary",    "Temporal",          "1 = transaction between 22:00 and 06:00",                           "is_night = 1 if hour >= 22 or hour < 6"),
    "is_weekend":            ("binary",    "Temporal",          "1 = Saturday or Sunday",                                            "is_weekend = 1 if day_of_week >= 5"),
    "is_business_hours":     ("binary",    "Temporal",          "1 = transaction between 09:00 and 17:00 on a weekday",              "is_business_hours = 1 if (9 <= hour < 17) and not is_weekend"),

    # ── Amount Derived ───────────────────────────────────────────────────────
    "amount_to_balance_ratio": ("float",  "Amount Derived",    "Fraction of average balance consumed by this transaction",           "amount_to_balance_ratio = amount / avg_balance"),
    "is_round_amount":         ("binary", "Amount Derived",    "1 = amount is a multiple of 1000",                                   "is_round_amount = 1 if amount % 1000 == 0"),
    "log_amount":              ("float",  "Amount Derived",    "Natural logarithm of amount; normalises heavy-tail distribution",    "log_amount = ln(amount)"),

    # ── Velocity (Account-Level Rolling) ─────────────────────────────────────
    "txn_count_last_1hr":       ("int",   "Account Velocity",  "Number of transactions from this sender account in last 1 hour",     "Rolling count over (account_id, timestamp) within 1h window"),
    "txn_count_last_24hr":      ("int",   "Account Velocity",  "Number of transactions in last 24 hours for this account",           "Rolling count within 24h window"),
    "total_amount_last_24hr":   ("float", "Account Velocity",  "Total transaction amount by this account in last 24 hours",          "Rolling sum(amount) within 24h window"),
    "total_amount_last_7d":     ("float", "Account Velocity",  "Total transaction amount by this account in last 7 days",            "Rolling sum(amount) within 7d window"),
    "total_amount_last_30d":    ("float", "Account Velocity",  "Total transaction amount by this account in last 30 days",           "Rolling sum(amount) within 30d window"),
    "dormancy_flag":            ("binary","Account Velocity",  "1 = account was inactive for ≥30 days before this transaction",      "dormancy_flag = 1 if days_since_last_txn >= 30"),
    "txn_velocity_cumulative":  ("int",   "Account Velocity",  "Cumulative transaction count for this account up to this event",     "Cumulative count of txns per account_id ordered by timestamp"),

    # ── Velocity (Customer-Level Rolling) ────────────────────────────────────
    "cust_txn_count_last_1hr":  ("int",   "Customer Velocity", "Customer-level transaction count in last 1 hour",                    "Rolling count over (customer_id, timestamp) within 1h window"),
    "cust_txn_count_last_24hr": ("int",   "Customer Velocity", "Customer-level transaction count in last 24 hours",                  "Rolling count within 24h window"),
    "cust_total_amount_last_24hr": ("float","Customer Velocity","Total transaction amount across all accounts of customer in last 24h","Rolling sum(amount) per customer within 24h window"),
    "cust_total_amount_last_30d":  ("float","Customer Velocity","Total transaction amount for customer in last 30 days",              "Rolling sum(amount) per customer within 30d window"),
    "cust_unique_accounts_30d": ("int",   "Customer Velocity", "Unique accounts used by the customer in last 30 days",               "Rolling nunique(account_id) per customer within 30d window"),

    # ── Rule-Based Flags ─────────────────────────────────────────────────────
    "rule_large_cash_deposit":       ("binary","Rule Flags","Cash deposit ≥ 50,000",                       "amount >= 50000 and cash_flag == 1 and debit_credit == 'credit'"),
    "rule_cash_just_below_threshold":("binary","Rule Flags","Cash transaction 9,000–10,000 (structuring probe)","8000 <= amount < 10000 and cash_flag == 1"),
    "rule_high_value_transfer":      ("binary","Rule Flags","Non-cash transfer ≥ 100,000",                  "amount >= 100000 and cash_flag == 0"),
    "rule_micro_transaction":        ("binary","Rule Flags","Amount < 10 (test transaction probe)",         "amount < 10"),
    "rule_round_amount_large":       ("binary","Rule Flags","Round amount ≥ 10,000",                        "is_round_amount == 1 and amount >= 10000"),
    "rule_night_transaction":        ("binary","Rule Flags","Transaction between 22:00–06:00",              "is_night == 1"),
    "rule_weekend_high_value":       ("binary","Rule Flags","Weekend transaction ≥ 50,000",                 "is_weekend == 1 and amount >= 50000"),
    "rule_dormant_account_activation":("binary","Rule Flags","Dormant account reactivated",                 "dormancy_flag == 1"),
    "rule_high_acct_velocity_1hr":   ("binary","Rule Flags","≥5 transactions from account in 1 hour",       "txn_count_last_1hr >= 5"),
    "rule_high_acct_velocity_24hr":  ("binary","Rule Flags","≥20 transactions from account in 24 hours",    "txn_count_last_24hr >= 20"),
    "rule_high_acct_velocity_7d":    ("binary","Rule Flags","≥100 transactions from account in 7 days",     "txn_count_last_7d >= 100"),
    "rule_high_acct_volume_24hr":    ("binary","Rule Flags","Account volume > 5× median in 24 hours",       "total_amount_last_24hr > 5 * median_daily_volume"),
    "rule_high_acct_volume_7d":      ("binary","Rule Flags","Account volume > 3× median in 7 days",         "total_amount_last_7d > 3 * median_7d_volume"),
    "rule_amount_spike_30d":         ("binary","Rule Flags","Single txn > 5× average 30d transaction",      "amount > 5 * (total_amount_last_30d / txn_count_30d)"),
    "rule_rapid_burst":              ("binary","Rule Flags","≥3 txns within 10 minutes",                    "txn_count_10min >= 3"),
    "rule_high_cust_velocity_1hr":   ("binary","Rule Flags","≥3 customer transactions in 1 hour",           "cust_txn_count_last_1hr >= 3"),
    "rule_high_cust_velocity_24hr":  ("binary","Rule Flags","≥10 customer transactions in 24 hours",        "cust_txn_count_last_24hr >= 10"),
    "rule_high_cust_volume_24hr":    ("binary","Rule Flags","Customer 24h volume > 3× account balance",     "cust_total_amount_last_24hr > 3 * avg_balance"),
    "rule_high_cust_volume_7d":      ("binary","Rule Flags","Customer 7d volume > 10× account balance",     "cust_total_amount_last_30d > 10 * avg_balance (7d proxy)"),
    "rule_low_kyc_high_amount":      ("binary","Rule Flags","KYC=low and amount ≥ 50,000",                  "kyc_level == 'low' and amount >= 50000"),
    "rule_new_account_large_txn":    ("binary","Rule Flags","Account < 30 days old and amount ≥ 10,000",    "account_age_days < 30 and amount >= 10000"),
    "rule_high_amount_to_balance":   ("binary","Rule Flags","Transaction > 200% of average balance",        "amount_to_balance_ratio > 2.0"),
    "rule_very_high_risk_customer":  ("binary","Rule Flags","Customer risk rating = very_high",             "customer_risk_rating == 'very_high'"),
    "rule_pep_high_value":           ("binary","Rule Flags","PEP customer and amount ≥ 10,000",             "pep_flag == 1 and amount >= 10000"),
    "rule_high_risk_country_sender": ("binary","Rule Flags","Sender's country risk = high",                 "country_risk == 'high'"),
    "rule_corporate_large_cash":     ("binary","Rule Flags","Corporate account and large cash transaction",  "account_type == 'corporate' and cash_flag == 1 and amount >= 50000"),
    "rule_high_risk_beneficiary":    ("binary","Rule Flags","Beneficiary is crypto or offshore",            "high_risk_beneficiary == 1"),
    "rule_crypto_transfer":          ("binary","Rule Flags","Transfer to a crypto beneficiary",             "beneficiary_type == 'crypto'"),
    "rule_offshore_transfer":        ("binary","Rule Flags","Transfer to an offshore beneficiary",          "beneficiary_type == 'offshore'"),
    "rule_high_risk_bene_country":   ("binary","Rule Flags","Beneficiary country risk = high",              "beneficiary_country_risk == 'high'"),
    "rule_rooted_device":            ("binary","Rule Flags","Device is rooted or jailbroken",               "rooted_flag == 1"),
    "rule_vpn_proxy_detected":       ("binary","Rule Flags","VPN or proxy active during transaction",       "vpn_flag == 1"),
    "rule_emulator_detected":        ("binary","Rule Flags","Emulator detected (potential bot)",            "emulator_flag == 1"),
    "rule_new_device_large_txn":     ("binary","Rule Flags","Device < 30 days old and amount ≥ 20,000",    "device_age_days < 30 and amount >= 20000"),
    "rule_atm_high_withdrawal":      ("binary","Rule Flags","ATM withdrawal ≥ 20,000",                     "channel == 'atm' and transaction_type == 'cash_withdrawal' and amount >= 20000"),
    "rule_branch_night_txn":         ("binary","Rule Flags","Branch transaction at night (22:00–06:00)",    "channel == 'branch' and is_night == 1"),
    "rule_structuring_pattern":      ("binary","Rule Flags","Multiple cash txns just below threshold (structuring)", "rule_cash_just_below_threshold == 1 and txn_count_last_24hr >= 3"),
    "rule_multiple_small_cash":      ("binary","Rule Flags","≥5 small cash transactions within 24 hours",  "cash_flag == 1 and amount < 5000 and txn_count_last_24hr >= 5"),
    "rule_high_risk_industry":       ("binary","Rule Flags","Customer operates in a high-risk industry",   "industry in {crypto, gambling, firearms, ...}"),
    "rule_student_high_value":       ("binary","Rule Flags","Student occupation and amount ≥ 50,000",       "occupation == 'student' and amount >= 50000"),
    "rule_unemployed_large_transfer":("binary","Rule Flags","Unemployed and amount ≥ 20,000",              "occupation == 'unemployed' and amount >= 20000"),
    "rule_freelancer_offshore":      ("binary","Rule Flags","Freelancer + offshore transfer",              "occupation == 'freelancer' and beneficiary_type == 'offshore'"),
    "rule_pep_crypto_transfer":      ("binary","Rule Flags","PEP customer making a crypto transfer",        "pep_flag == 1 and beneficiary_type == 'crypto'"),
    "rule_very_high_risk_offshore":  ("binary","Rule Flags","Very-high-risk customer + offshore transfer", "customer_risk_rating == 'very_high' and beneficiary_type == 'offshore'"),
    "rule_low_kyc_offshore":         ("binary","Rule Flags","KYC=low + offshore transfer",                 "kyc_level == 'low' and beneficiary_type == 'offshore'"),
    "rule_low_income_large_txn":     ("binary","Rule Flags","Low-income bracket and amount ≥ 50,000",      "income_bracket == 'low' and amount >= 50000"),
    "rule_vpn_offshore":             ("binary","Rule Flags","VPN active + offshore transfer",              "vpn_flag == 1 and beneficiary_type == 'offshore'"),
    "rule_emulator_crypto":          ("binary","Rule Flags","Emulator + crypto transfer",                  "emulator_flag == 1 and beneficiary_type == 'crypto'"),
    "rule_new_account_offshore":     ("binary","Rule Flags","New account (< 30 days) + offshore transfer", "account_age_days < 30 and beneficiary_type == 'offshore'"),
    "rule_new_acct_high_cust_velocity":("binary","Rule Flags","New account + high customer velocity",     "account_age_days < 30 and cust_txn_count_last_24hr >= 5"),
    "rule_trigger_count":            ("int",   "Rule Aggregates","Total number of rules triggered for this transaction", "rule_trigger_count = sum(all rule_* columns)"),
    "max_rule_severity":             ("int",   "Rule Aggregates","Severity level of the most severe triggered rule (0–3)", "max_rule_severity = max severity tier across triggered rules"),
    "weighted_rule_score":           ("int",   "Rule Aggregates","Weighted sum of triggered rule severities", "Σ (rule_i × severity_weight_i) for all triggered rules"),

    # ── Graph Features ───────────────────────────────────────────────────────
    "sender_out_degree_30d":         ("float","Graph – Sender",   "Number of outgoing transactions from sender account in last 30 days", "Rolling count(txn_id) per sender grouped by sender_account_id, window=30d"),
    "sender_total_outflow_30d":      ("float","Graph – Sender",   "Total amount sent from sender account in last 30 days",              "Rolling sum(amount) per sender, window=30d"),
    "sender_unique_counterparties_30d":("int","Graph – Sender",   "Distinct receiver accounts/beneficiaries contacted in last 30 days", "Rolling nunique(receiver) per sender, window=30d"),
    "sender_repeat_counterparty_ratio":("float","Graph – Sender", "Fraction of outgoing transactions going to already-seen counterparties","(out_degree_30d – unique_counterparties_30d) / out_degree_30d"),
    "sender_in_degree_30d":          ("float","Graph – Sender",   "Number of incoming transactions to sender account in last 30 days (sparse)", "Rolling count per account from inflow side; sparse – only 49 rows populated"),
    "sender_total_inflow_30d":       ("float","Graph – Sender",   "Total amount received by sender account in last 30 days (sparse)",   "Rolling sum(amount) from inflow, window=30d; sparse"),
    "receiver_in_degree_30d":        ("float","Graph – Receiver", "Number of incoming transactions to the receiver account in last 30 days","Rolling count per receiver_account_id, window=30d; 0 for external transfers"),
    "receiver_total_inflow_30d":     ("float","Graph – Receiver", "Total amount received by the receiver account in last 30 days",       "Rolling sum(amount) per receiver, window=30d; 0 for external"),
    "receiver_unique_senders_30d":   ("float","Graph – Receiver", "Distinct senders who sent money to receiver account in last 30 days", "Rolling nunique(sender) per receiver, window=30d; 0 for external"),
    "receiver_account_outflow_30d":  ("float","Graph – Receiver", "Number of outgoing transactions from the receiver account in last 30 days","Reuse of sender_out_degree_30d joined on receiver_account_id; 0 for external"),
    "inflow_outflow_volume_balance_ratio_24h":("float","Graph – Layering","Volume balance ratio: min(inflow,outflow)/inflow over 24 hours","min(sender_total_inflow_24h, sender_total_outflow_24h) / sender_total_inflow_24h; 0 if no inflow"),
    "inflow_outflow_volume_balance_ratio_7d": ("float","Graph – Layering","Same ratio computed over a 7-day rolling window",             "min(sender_total_inflow_7d, sender_total_outflow_7d) / sender_total_inflow_7d"),
    "outflow_to_inflow_ratio_7d":    ("float","Graph – Layering","Ratio of total outflow to total inflow over 7 days",                  "sender_total_outflow_7d / sender_total_inflow_7d; 0 if no inflow"),
    "devices_per_account":           ("int",  "Graph – Device",  "Number of distinct devices linked to this sender account in last 30 days","Rolling nunique(device_id) per account, window=30d"),
    "accounts_per_device":           ("int",  "Graph – Device",  "Number of distinct accounts that used this device in last 30 days",    "Rolling nunique(account_id) per device, window=30d"),
    "device_shared_high_risk_ratio": ("float","Graph – Device",  "Fraction of device's recent transactions that involved a fraud or high-risk beneficiary","Rolling sum(_fraud_or_high_risk) / count per device, window=30d"),
    "shared_device_fraud_count":     ("float","Graph – Device",  "Rolling count of fraud or high-risk transactions on the same device in last 30 days","Rolling sum(label==1 or high_risk_beneficiary==1) per device, window=30d"),
    "avg_time_gap_in_out":           ("float","Graph – Timing",  "Rolling average time gap (seconds) between the latest inflow and next outflow for sender account","Per account, for each inflow event, find next outflow and compute gap_sec; rolling 24h mean"),
    "account_id_x":                  ("string","Graph – Internal","Internal join key (sender side); not a model feature",               "Intermediate artifact from merge; can be dropped"),
    "account_id_y":                  ("string","Graph – Internal","Internal join key (receiver side); not a model feature",             "Intermediate artifact from merge; can be dropped"),
}

def sheet_data_dict(wb, df):
    ws = wb.create_sheet("2_Data_Dictionary")
    ws.freeze_panes = "A3"

    title_cell(ws, 1, 1, "Data Dictionary – All Features & Definitions")
    headers = ["Column Name","Data Type","Feature Group","Definition","Formula / Derivation"]
    apply_header(ws, 2, headers, bg=C_DARK_BLUE)

    group_colors = {
        "Transaction Base": "E8F5E9", "Labels": "FCE4D6",
        "Customer Profile": "E3F2FD", "Device": "FFF9C4",
        "Beneficiary": "F3E5F5", "Temporal": "E0F7FA",
        "Amount Derived": "FFF3E0", "Account Velocity": "E8EAF6",
        "Customer Velocity": "E8EAF6", "Rule Flags": "FFF8E1",
        "Rule Aggregates": "FFF8E1",
        "Graph – Sender": "E8F5E9", "Graph – Receiver": "E8F5E9",
        "Graph – Layering": "FCE4D6", "Graph – Device": "E3F2FD",
        "Graph – Timing": "FFF3E0", "Graph – Internal": "F2F2F2",
    }

    row = 3
    for col_name in df.columns:
        if col_name in DATA_DICT:
            dtype, group, defn, formula = DATA_DICT[col_name]
        else:
            dtype = str(df[col_name].dtype)
            group = "Other"
            defn  = ""
            formula = ""

        bg = group_colors.get(group, "FFFFFF")
        vals = [col_name, dtype, group, defn, formula]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = body_font(bold=(ci == 1))
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1

    set_col_widths(ws, [(1,28),(2,14),(3,22),(4,60),(5,70)])
    ws.row_dimensions[1].height = 20
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 2_Data_Dictionary")


# =============================================================================
# SHEET 3 – FEATURE FORMULAS
# =============================================================================
FORMULAS = [
    # category, feature, mathematical_formula, aml_intuition, detection_target
    ("Amount Derived", "amount_to_balance_ratio",
     "amount / avg_balance",
     "Measures how large this transaction is relative to the customer's typical holding. Values > 1 mean the customer is sending more than their average balance.",
     "ATO (account takeover), Mule Ring (large drain)"),

    ("Amount Derived", "log_amount",
     "ln(amount)",
     "Natural log transform; compresses the heavy-tailed distribution of transaction amounts so ML models receive a more Gaussian-like input.",
     "All fraud types (feature preprocessing)"),

    ("Temporal", "is_night",
     "1 if hour ∈ [22,23,0,1,2,3,4,5], else 0",
     "Fraudsters operate at night to avoid human oversight and real-time fraud monitoring.",
     "Layering, ATO, Smurfing"),

    ("Temporal", "is_weekend",
     "1 if day_of_week ∈ {5, 6}, else 0",
     "Weekend transactions receive less scrutiny from operations teams.",
     "Layering, ATO"),

    ("Temporal", "is_business_hours",
     "1 if 9 ≤ hour < 17 and day_of_week ∈ {0..4}, else 0",
     "Negative of is_night; legitimate corporate transactions are concentrated in business hours.",
     "All (contrast feature)"),

    ("Account Velocity", "txn_count_last_Xhr",
     "Rolling count of transactions per sender_account_id within X-hour window ending at timestamp",
     "Rapid transaction bursts indicate automated/scripted activity (bots, mule accounts).",
     "Mule Ring, Smurfing, ATO"),

    ("Account Velocity", "total_amount_last_Xd",
     "Rolling sum(amount) per sender_account_id within X-day window ending at timestamp",
     "Accumulation of large volumes over short periods signals layering or rapid fund movement.",
     "Layering, Mule Ring"),

    ("Account Velocity", "dormancy_flag",
     "1 if days since last transaction ≥ 30, else 0",
     "Dormant accounts re-activated with large transactions are a classic ATO or mule pattern.",
     "ATO, Mule Ring"),

    ("Account Velocity", "txn_velocity_cumulative",
     "Cumulative count of transactions for sender_account_id ordered by timestamp",
     "Very young accounts (low cumulative count) making large transactions are suspicious.",
     "Identity Fraud, ATO"),

    ("Customer Velocity", "cust_unique_accounts_30d",
     "Rolling nunique(account_id) per customer_id within 30d window",
     "Using many different accounts suggests identity fraud or account takeover of multiple victims.",
     "Identity Fraud, Mule Ring"),

    ("Graph – Sender", "sender_out_degree_30d",
     "Rolling COUNT(transaction_id) per sender_account_id, window = 30 days",
     "High fan-out (many outgoing transactions) from a single account indicates mule or layering activity.",
     "Mule Ring, Layering"),

    ("Graph – Sender", "sender_unique_counterparties_30d",
     "Rolling NUNIQUE(receiver_account_id ∪ beneficiary_id) per sender, window = 30d",
     "Sending to many distinct recipients indicates wide dispersion (characteristic of smurfing/layering).",
     "Layering, Smurfing"),

    ("Graph – Sender", "sender_repeat_counterparty_ratio",
     "(sender_out_degree_30d − sender_unique_counterparties_30d) / sender_out_degree_30d",
     "Low ratio = many new counterparties (dispersion). High ratio = repeat contacts (salary, recurring bills). Fraud tends toward low ratio.",
     "Layering, Smurfing"),

    ("Graph – Receiver", "receiver_in_degree_30d",
     "Rolling COUNT(transaction_id) per receiver_account_id, window = 30d",
     "Accounts that receive from many senders are potential collection points (smurfing aggregation).",
     "Smurfing, Mule Ring"),

    ("Graph – Receiver", "receiver_account_outflow_30d",
     "Reuse of sender_out_degree_30d joined on receiver_account_id",
     "A receiver that also has high outflow is acting as a pass-through (mule). External transfers default to 0.",
     "Mule Ring, Layering"),

    ("Graph – Layering", "inflow_outflow_volume_balance_ratio_24h",
     "min(sender_total_inflow_24h, sender_total_outflow_24h) / sender_total_inflow_24h\nwhere 24h inflow/outflow = rolling sums over sender_account_id",
     "Ratio → 1 means all inflow is immediately re-transmitted (pass-through). Pure layering nodes score near 1. External-transfer accounts score 0.",
     "Layering, Mule Ring"),

    ("Graph – Layering", "inflow_outflow_volume_balance_ratio_7d",
     "min(sender_total_inflow_7d, sender_total_outflow_7d) / sender_total_inflow_7d",
     "7-day version provides a slower-burn view of pass-through behaviour vs. short-term spikes.",
     "Layering"),

    ("Graph – Layering", "outflow_to_inflow_ratio_7d",
     "sender_total_outflow_7d / sender_total_inflow_7d",
     "Ratios >> 1 mean the account is spending more than received (funded externally / ATO). Ratios ≈ 1 confirm pass-through.",
     "ATO, Layering"),

    ("Graph – Device", "devices_per_account",
     "Rolling NUNIQUE(device_id) per sender_account_id, window = 30d",
     "Legitimate customers use 1–2 devices. ATO victims show sudden new device. Mules show many devices.",
     "ATO, Mule Ring"),

    ("Graph – Device", "accounts_per_device",
     "Rolling NUNIQUE(account_id) per device_id, window = 30d",
     "A device shared across many accounts indicates a device farm or coordinated mule ring.",
     "Mule Ring, Identity Fraud"),

    ("Graph – Device", "device_shared_high_risk_ratio",
     "Rolling SUM(_fraud_or_high_risk) / COUNT(*) per device_id, window = 30d\nwhere _fraud_or_high_risk = (label==1) | (high_risk_beneficiary==1)",
     "If many transactions on a shared device are high-risk, the device is part of a fraud ring. Avoids double-counting by using union flag.",
     "Mule Ring, ATO"),

    ("Graph – Device", "shared_device_fraud_count",
     "Rolling SUM(_fraud_or_high_risk) per device_id, window = 30d",
     "Raw count version of device_shared_high_risk_ratio; useful as an absolute signal alongside the ratio.",
     "Mule Ring, ATO"),

    ("Graph – Timing", "avg_time_gap_in_out",
     "For each inflow event of sender_account_id at time t_in:\n  1. Find first outflow at t_out > t_in\n  2. gap_sec = (t_out − t_in).seconds\nThen rolling mean of gap_sec over 24h window\nMerged to main df via per-account as-of join (np.searchsorted)",
     "Short average gaps (seconds to minutes) between receiving funds and forwarding them signals layering / mule pass-through. Long gaps indicate normal usage.",
     "Layering, Mule Ring"),
]

def sheet_formulas(wb):
    ws = wb.create_sheet("3_Feature_Formulas")
    ws.freeze_panes = "A3"

    title_cell(ws, 1, 1, "Feature Formulas, Derivations & AML Intuition")
    headers = ["Feature Group","Feature Name","Mathematical Formula / Derivation","AML Intuition","Fraud Types Targeted"]
    apply_header(ws, 2, headers, bg=C_DARK_BLUE)

    group_colors = {
        "Amount Derived": "FFF3E0", "Temporal": "E0F7FA",
        "Account Velocity": "E8EAF6", "Customer Velocity": "E8EAF6",
        "Graph – Sender": "E8F5E9", "Graph – Receiver": "E8F5E9",
        "Graph – Layering": "FCE4D6", "Graph – Device": "E3F2FD",
        "Graph – Timing": "FFF9C4",
    }

    row = 3
    for group, feat, formula, intuition, target in FORMULAS:
        bg = group_colors.get(group, C_WHITE)
        vals = [group, feat, formula, intuition, target]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = body_font(bold=(ci <= 2))
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 60
        row += 1

    set_col_widths(ws, [(1,22),(2,35),(3,60),(4,70),(5,35)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 3_Feature_Formulas")


# =============================================================================
# SHEET 4 – DESCRIPTIVE STATISTICS
# =============================================================================
def sheet_desc_stats(wb, df):
    ws = wb.create_sheet("4_Descriptive_Stats")
    ws.freeze_panes = "A3"

    title_cell(ws, 1, 1, "Descriptive Statistics – All Numeric Features")

    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    numeric_cols = [c for c in numeric_cols if c not in ["account_id_x","account_id_y"]]

    stats = df[numeric_cols].describe(percentiles=[.25,.5,.75,.90,.95,.99]).T.round(4)
    stats.insert(0, "feature", stats.index)
    stats = stats.reset_index(drop=True)

    # Add null count and null %
    stats["null_count"] = df[numeric_cols].isnull().sum().values
    stats["null_pct"]   = (df[numeric_cols].isnull().mean() * 100).round(2).values

    headers = list(stats.columns)
    apply_header(ws, 2, headers, bg=C_DARK_BLUE)

    for r_i, row_data in enumerate(stats.itertuples(index=False), start=3):
        bg = C_LIGHT_BLUE if r_i % 2 == 0 else C_WHITE
        for c_i, val in enumerate(row_data, 1):
            c = ws.cell(row=r_i, column=c_i, value=val)
            c.font = body_font(bold=(c_i == 1))
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()

    # Auto-width
    for ci in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions["A"].width = 40
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 4_Descriptive_Stats")


# =============================================================================
# SHEET 5 – MISSING VALUES
# =============================================================================
def sheet_missing(wb, df):
    ws = wb.create_sheet("5_Missing_Values")
    ws.freeze_panes = "A3"

    title_cell(ws, 1, 1, "Missing Value Analysis")
    headers = ["Column","Data Type","Null Count","Null %","Feature Group","Notes"]
    apply_header(ws, 2, headers, bg=C_DARK_BLUE)

    nulls = df.isnull().sum().reset_index()
    nulls.columns = ["column","null_count"]
    nulls["null_pct"] = (nulls["null_count"] / len(df) * 100).round(2)
    nulls["dtype"] = nulls["column"].map(lambda c: str(df[c].dtype))
    nulls["group"] = nulls["column"].map(lambda c: DATA_DICT.get(c, ("","","",""))[1])
    nulls = nulls.sort_values("null_count", ascending=False)

    notes_map = {
        "receiver_account_id":   "NULL when transfer goes to external beneficiary (~62% of rows). Expected.",
        "beneficiary_id":        "NULL for intra-bank transfers. Expected.",
        "beneficiary_type":      "NULL for intra-bank transfers. Expected.",
        "beneficiary_country_risk": "NULL for intra-bank transfers. Expected.",
        "account_id_x":          "Internal join artifact from graph feature merge. Not used in modelling.",
        "sender_in_degree_30d":  "Internal join artifact; only 49 rows populated. Not used in modelling.",
        "sender_total_inflow_30d":"Internal join artifact; only 49 rows populated. Not used in modelling.",
        "account_id_y":          "Internal join artifact from graph feature merge. Not used in modelling.",
        "avg_time_gap_in_out":   "NULL for accounts that have never had a subsequent outflow after an inflow (~4.5% of rows). Can be imputed with median or 0.",
    }

    row = 3
    for _, r in nulls.iterrows():
        count = r["null_count"]
        pct   = r["null_pct"]
        if count == 0:
            bg = C_WHITE
        elif pct > 50:
            bg = "FCE4D6"
        elif pct > 5:
            bg = "FFF9C4"
        else:
            bg = "FFFDE7"

        note = notes_map.get(r["column"], "")
        vals = [r["column"], r["dtype"], count, f"{pct}%", r["group"], note]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    set_col_widths(ws, [(1,30),(2,14),(3,14),(4,12),(5,22),(6,65)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 5_Missing_Values")


# =============================================================================
# SHEET 6 – GRAPH FEATURES EDA
# =============================================================================
def sheet_graph_eda(wb, df):
    ws = wb.create_sheet("6_Graph_Features_EDA")
    ws.freeze_panes = "A4"

    title_cell(ws, 1, 1, "Graph Features – EDA & Fraud Signal Analysis")

    graph_cols = [
        "sender_out_degree_30d","sender_total_outflow_30d","sender_unique_counterparties_30d",
        "sender_repeat_counterparty_ratio","sender_in_degree_30d","sender_total_inflow_30d",
        "receiver_in_degree_30d","receiver_total_inflow_30d","receiver_unique_senders_30d",
        "receiver_account_outflow_30d","inflow_outflow_volume_balance_ratio_24h",
        "inflow_outflow_volume_balance_ratio_7d","outflow_to_inflow_ratio_7d",
        "devices_per_account","accounts_per_device","device_shared_high_risk_ratio",
        "shared_device_fraud_count","avg_time_gap_in_out"
    ]

    # ── Section 1: Descriptive stats ─────────────────────────────────────────
    section_cell(ws, 3, 1, "Section 1 – Descriptive Statistics", bg=C_MED_BLUE)
    hdrs = ["Feature","Non-Null Count","Mean","Std Dev","Min","25%","Median","75%","90%","99%","Max","Null %"]
    apply_header(ws, 4, hdrs, bg=C_DARK_BLUE)
    row = 5
    for col in graph_cols:
        s = df[col].dropna()
        null_pct = df[col].isnull().mean()*100
        vals = [col, len(s), round(s.mean(),4), round(s.std(),4),
                round(s.min(),4), round(s.quantile(.25),4), round(s.median(),4),
                round(s.quantile(.75),4), round(s.quantile(.90),4), round(s.quantile(.99),4),
                round(s.max(),4), f"{null_pct:.1f}%"]
        bg = C_LIGHT_BLUE if row%2==0 else C_WHITE
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = body_font(bold=(ci==1))
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    # ── Section 2: Fraud vs Legit comparison ─────────────────────────────────
    row += 2
    section_cell(ws, row, 1, "Section 2 – Fraud vs Legitimate Mean Comparison", bg=C_ORANGE)
    row += 1
    hdrs2 = ["Feature","Legit Mean","Fraud Mean","Fraud/Legit Ratio","Signal Direction","AML Pattern"]
    apply_header(ws, row, hdrs2, bg=C_DARK_BLUE)
    row += 1
    aml_signals = {
        "sender_out_degree_30d":              ("↑ fraud higher", "Mule Ring: high fan-out from account"),
        "sender_total_outflow_30d":           ("↑ fraud higher", "Layering: large volumes moved"),
        "sender_unique_counterparties_30d":   ("↑ fraud higher", "Smurfing/Layering: wide dispersion"),
        "sender_repeat_counterparty_ratio":   ("↑ fraud higher", "Repeat contacts (mule ring coordination)"),
        "sender_in_degree_30d":               ("↑ fraud higher", "Collection point for incoming funds"),
        "sender_total_inflow_30d":            ("↑ fraud higher", "High inflow before redistribution"),
        "receiver_in_degree_30d":             ("↓ fraud lower",  "Fraud targets new/clean receivers"),
        "receiver_total_inflow_30d":          ("↓ fraud lower",  "Fraud targets low-activity receivers"),
        "receiver_unique_senders_30d":        ("↓ fraud lower",  "Fraud receivers have few prior senders"),
        "receiver_account_outflow_30d":       ("↑ fraud higher", "Mule receivers also have high outflow (pass-through)"),
        "inflow_outflow_volume_balance_ratio_24h":("↑ fraud higher","Layering: inflow quickly re-transmitted"),
        "inflow_outflow_volume_balance_ratio_7d": ("↑ fraud higher","Layering: persistent pass-through"),
        "outflow_to_inflow_ratio_7d":         ("↑ fraud higher", "ATO/Layering: spending more than receiving"),
        "devices_per_account":                ("↑ fraud higher", "ATO: new device on victim account"),
        "accounts_per_device":                ("↑ fraud higher", "Mule Ring: device farm shared by many accounts"),
        "device_shared_high_risk_ratio":      ("↑ fraud higher", "Mule Ring: device used in coordinated fraud"),
        "shared_device_fraud_count":          ("↑ fraud higher", "Mule Ring: device with many prior fraud events"),
        "avg_time_gap_in_out":                ("↓ fraud lower",  "Layering: rapid pass-through (short gap)"),
    }
    for col in graph_cols:
        non_null = df[col].notna()
        if non_null.sum() == 0:
            continue
        legit_m = df.loc[(non_null) & (df["label"]==0), col].mean()
        fraud_m = df.loc[(non_null) & (df["label"]==1), col].mean()
        ratio   = round(fraud_m/legit_m, 2) if legit_m else float("nan")
        direction, pattern = aml_signals.get(col, ("–","–"))

        # Highlight strong signals
        if abs(ratio) > 3:
            bg = "FCE4D6"
        elif abs(ratio) > 1.5:
            bg = "FFF9C4"
        else:
            bg = C_WHITE

        vals = [col, round(legit_m,4), round(fraud_m,4), f"{ratio:.2f}x", direction, pattern]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = body_font(bold=(ci==1))
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    set_col_widths(ws, [(1,42),(2,16),(3,16),(4,18),(5,18),(6,16),(7,18),(8,18),(9,18),(10,16),(11,16),(12,12)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 6_Graph_Features_EDA")


# =============================================================================
# SHEET 7 – CATEGORICAL DISTRIBUTIONS
# =============================================================================
def sheet_categoricals(wb, df):
    ws = wb.create_sheet("7_Categorical_Distributions")
    title_cell(ws, 1, 1, "Categorical Feature Distributions & Fraud Rates")

    cat_cols = [
        ("channel",                 "Channel"),
        ("transaction_type",        "Transaction Type"),
        ("kyc_level",               "KYC Level"),
        ("country_risk",            "Country Risk"),
        ("income_bracket",          "Income Bracket"),
        ("customer_risk_rating",    "Customer Risk Rating"),
        ("occupation",              "Occupation"),
        ("account_type",            "Account Type"),
        ("os_type",                 "OS Type"),
        ("beneficiary_type",        "Beneficiary Type"),
        ("beneficiary_country_risk","Beneficiary Country Risk"),
        ("debit_credit",            "Debit / Credit"),
    ]

    row = 3
    for col, label in cat_cols:
        if col not in df.columns:
            continue
        section_cell(ws, row, 1, f"{label}  [{col}]", bg=C_MED_BLUE)
        for ci in range(2, 6):
            ws.cell(row=row, column=ci).fill = fill(C_MED_BLUE)
            ws.cell(row=row, column=ci).border = thin_border()
        row += 1
        apply_header(ws, row, ["Value","Total Count","% of All","Fraud Count","Fraud Rate (%)"], bg=C_DARK_BLUE)
        row += 1
        tbl = df.groupby(col, dropna=False)["label"].agg(["count","sum","mean"]).reset_index()
        tbl.columns = [col,"total","fraud","fraud_rate"]
        tbl["pct_all"] = (tbl["total"] / len(df) * 100).round(2)
        tbl = tbl.sort_values("total", ascending=False)

        for _, r in tbl.iterrows():
            bg = "FCE4D6" if r["fraud_rate"] > 0.4 else (C_LIGHT_BLUE if row%2==0 else C_WHITE)
            vals = [r[col], r["total"], f"{r['pct_all']}%", r["fraud"], f"{r['fraud_rate']*100:.1f}%"]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = body_font()
                c.fill = fill(bg)
                c.border = thin_border()
                c.alignment = left()
            row += 1
        row += 2

    set_col_widths(ws, [(1,35),(2,14),(3,12),(4,14),(5,15)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 7_Categorical_Distributions")


# =============================================================================
# SHEET 8 – RULE FEATURES
# =============================================================================
def sheet_rules(wb, df):
    ws = wb.create_sheet("8_Rule_Features")
    ws.freeze_panes = "A4"

    title_cell(ws, 1, 1, "Rule-Based Feature Analysis – Trigger Rates & Fraud Correlation")

    rule_cols = [c for c in df.columns if c.startswith("rule_") and c not in ["rule_trigger_count"]]

    # ── Section 1: Rule Stats ─────────────────────────────────────────────────
    section_cell(ws, 3, 1, "Section 1 – Per-Rule Statistics", bg=C_MED_BLUE)
    hdrs = ["Rule Name","Trigger Count","Trigger Rate (%)","Fraud Rate When Triggered (%)","Fraud Rate When NOT Triggered (%)","Lift (ratio)","Signal Strength"]
    apply_header(ws, 4, hdrs, bg=C_DARK_BLUE)
    row = 5

    rule_stats = []
    for c in rule_cols:
        total_triggered = int(df[c].sum())
        trigger_rate    = df[c].mean() * 100
        f_triggered     = df.loc[df[c]==1, "label"].mean() * 100 if total_triggered > 0 else 0
        f_not_triggered = df.loc[df[c]==0, "label"].mean() * 100
        lift = f_triggered / f_not_triggered if f_not_triggered > 0 else float("nan")
        rule_stats.append((c, total_triggered, trigger_rate, f_triggered, f_not_triggered, lift))

    rule_stats.sort(key=lambda x: x[3], reverse=True)

    for (c, cnt, tr, ft, fn, lift) in rule_stats:
        if ft > 50:
            bg, strength = "FCE4D6", "🔴 Very High"
        elif ft > 30:
            bg, strength = "FFF3E0", "🟠 High"
        elif ft > 15:
            bg, strength = "FFF9C4", "🟡 Medium"
        else:
            bg, strength = C_WHITE, "⚪ Low"

        vals = [c, cnt, f"{tr:.2f}%", f"{ft:.1f}%", f"{fn:.1f}%", f"{lift:.2f}x" if not np.isnan(lift) else "N/A", strength]
        for ci, v in enumerate(vals, 1):
            c_ = ws.cell(row=row, column=ci, value=v)
            c_.font = body_font()
            c_.fill = fill(bg)
            c_.border = thin_border()
            c_.alignment = left()
        row += 1

    # ── Section 2: Trigger count distribution ───────────────────────────────
    row += 2
    section_cell(ws, row, 1, "Section 2 – Rule Trigger Count Distribution", bg=C_MED_BLUE)
    row += 1
    apply_header(ws, row, ["Trigger Count","# Transactions","% of All","Fraud Count","Fraud Rate (%)"], bg=C_DARK_BLUE)
    row += 1
    tc = df.groupby("rule_trigger_count")["label"].agg(["count","sum","mean"]).reset_index()
    tc.columns = ["trigger_count","total","fraud","fraud_rate"]
    for _, r in tc.iterrows():
        bg = "FCE4D6" if r["fraud_rate"] > 0.5 else (C_LIGHT_BLUE if row%2==0 else C_WHITE)
        vals = [r["trigger_count"], r["total"], f"{r['total']/len(df)*100:.1f}%", r["fraud"], f"{r['fraud_rate']*100:.1f}%"]
        for ci, v in enumerate(vals, 1):
            c_ = ws.cell(row=row, column=ci, value=v)
            c_.font = body_font()
            c_.fill = fill(bg)
            c_.border = thin_border()
            c_.alignment = left()
        row += 1

    # ── Section 3: Max severity distribution ─────────────────────────────────
    row += 2
    section_cell(ws, row, 1, "Section 3 – Max Rule Severity vs Fraud Rate", bg=C_MED_BLUE)
    row += 1
    apply_header(ws, row, ["Severity Level","Description","# Transactions","Fraud Count","Fraud Rate (%)"], bg=C_DARK_BLUE)
    row += 1
    sev_desc = {0:"No rules triggered",1:"Low-severity rule(s) only",2:"Medium-severity rule(s)",3:"High-severity rule(s)"}
    ms = df.groupby("max_rule_severity")["label"].agg(["count","sum","mean"]).reset_index()
    for _, r in ms.iterrows():
        bg = "FCE4D6" if r["mean"] > 0.3 else (C_LIGHT_BLUE if row%2==0 else C_WHITE)
        vals = [int(r["max_rule_severity"]), sev_desc.get(int(r["max_rule_severity"]),""), r["count"], r["sum"], f"{r['mean']*100:.1f}%"]
        for ci, v in enumerate(vals, 1):
            c_ = ws.cell(row=row, column=ci, value=v)
            c_.font = body_font()
            c_.fill = fill(bg)
            c_.border = thin_border()
            c_.alignment = left()
        row += 1

    set_col_widths(ws, [(1,42),(2,16),(3,18),(4,28),(5,28),(6,14),(7,16)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 8_Rule_Features")


# =============================================================================
# SHEET 9 – TEMPORAL ANALYSIS
# =============================================================================
def sheet_temporal(wb, df):
    ws = wb.create_sheet("9_Temporal_Analysis")
    title_cell(ws, 1, 1, "Temporal Analysis – Fraud Patterns Over Time")

    # Hour
    row = 3
    section_cell(ws, row, 1, "Fraud Rate by Hour of Day", bg=C_MED_BLUE)
    row += 1
    apply_header(ws, row, ["Hour","Total Txns","Fraud Count","Fraud Rate (%)","Period"], bg=C_DARK_BLUE)
    row += 1
    hour_tbl = df.groupby("hour")["label"].agg(["count","sum","mean"]).reset_index()
    for _, r in hour_tbl.iterrows():
        h = int(r["hour"])
        period = "Night (22-05)" if (h >= 22 or h < 6) else ("Early Morning (6-8)" if h < 9 else ("Business (9-17)" if h < 17 else "Evening (17-21)"))
        bg = "FCE4D6" if r["mean"] > 0.3 else (C_LIGHT_BLUE if row%2==0 else C_WHITE)
        vals = [h, r["count"], r["sum"], f"{r['mean']*100:.1f}%", period]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    # Day of week
    row += 2
    section_cell(ws, row, 1, "Fraud Rate by Day of Week", bg=C_MED_BLUE)
    row += 1
    apply_header(ws, row, ["Day of Week (0=Mon)","Day Name","Total","Fraud Count","Fraud Rate (%)"], bg=C_DARK_BLUE)
    row += 1
    days = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    dow = df.groupby("day_of_week")["label"].agg(["count","sum","mean"]).reset_index()
    for _, r in dow.iterrows():
        bg = C_LIGHT_BLUE if row%2==0 else C_WHITE
        vals = [int(r["day_of_week"]), days[int(r["day_of_week"])], r["count"], r["sum"], f"{r['mean']*100:.1f}%"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    # Month
    row += 2
    section_cell(ws, row, 1, "Fraud Rate by Month", bg=C_MED_BLUE)
    row += 1
    apply_header(ws, row, ["Month","Month Name","Total","Fraud Count","Fraud Rate (%)"], bg=C_DARK_BLUE)
    row += 1
    months = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    mo = df.groupby("month")["label"].agg(["count","sum","mean"]).reset_index()
    for _, r in mo.iterrows():
        bg = C_LIGHT_BLUE if row%2==0 else C_WHITE
        vals = [int(r["month"]), months[int(r["month"])], r["count"], r["sum"], f"{r['mean']*100:.1f}%"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = body_font()
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    set_col_widths(ws, [(1,22),(2,18),(3,14),(4,14),(5,16)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 9_Temporal_Analysis")


# =============================================================================
# SHEET 10 – FRAUD TYPE ANALYSIS
# =============================================================================
def sheet_fraud_types(wb, df):
    ws = wb.create_sheet("10_Fraud_Type_Analysis")
    title_cell(ws, 1, 1, "Fraud Type Deep-Dive Analysis")

    fraud_types = ["mule_ring","layering","smurfing","ATO","identity_fraud"]
    rule_cols = [c for c in df.columns if c.startswith("rule_") and c not in ["rule_trigger_count"]]

    fraud_descriptions = {
        "mule_ring": "Money Mule Ring – Recruited individuals receive and forward funds to obscure origin. Characterized by high-risk beneficiaries, shared devices, UPI/wallet transfers.",
        "layering": "Layering – Rapid movement of funds through multiple accounts to obscure origin. Characterized by night transactions, high pass-through ratios, multiple hops.",
        "smurfing": "Smurfing (Structuring) – Breaking large amounts into smaller transactions below reporting thresholds. Characterized by cash transactions just below 10,000, structuring rule triggers.",
        "ATO": "Account Takeover – Fraudster gains access to legitimate account and drains it. Characterized by high-value RTGS/web transactions, new device usage, off-hours activity.",
        "identity_fraud": "Identity Fraud – Using stolen/fabricated identity to open and exploit accounts. Characterized by new accounts, offshore transfers, UPI.",
    }

    row = 3
    for ft in fraud_types:
        sub = df[df["fraud_type"] == ft]
        normal = df[df["fraud_type"] == "normal"]

        section_cell(ws, row, 1, f"■  {ft.upper()}  (n = {len(sub):,})", bg=C_DARK_BLUE)
        for ci in range(2, 7):
            ws.cell(row=row, column=ci).fill = fill(C_DARK_BLUE)
            ws.cell(row=row, column=ci).border = thin_border()
        row += 1

        # Description
        c = ws.cell(row=row, column=1, value=fraud_descriptions.get(ft,""))
        c.font = Font(italic=True, size=10, name="Calibri", color="333333")
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 40
        row += 1

        # Key stats
        apply_header(ws, row, ["Metric","Fraud Value","Normal Value","Ratio"], bg=C_MED_BLUE)
        row += 1
        key_metrics = [
            ("Transaction Count",        len(sub),                           len(normal),                      ""),
            ("Average Amount",           f"{sub['amount'].mean():.2f}",      f"{normal['amount'].mean():.2f}", f"{sub['amount'].mean()/normal['amount'].mean():.2f}x"),
            ("Median Amount",            f"{sub['amount'].median():.2f}",    f"{normal['amount'].median():.2f}",""),
            ("Night Txn Rate",           f"{sub['is_night'].mean()*100:.1f}%", f"{normal['is_night'].mean()*100:.1f}%",""),
            ("Avg devices_per_account",  f"{sub['devices_per_account'].mean():.2f}", f"{normal['devices_per_account'].mean():.2f}", f"{sub['devices_per_account'].mean()/normal['devices_per_account'].mean():.2f}x"),
            ("Avg accounts_per_device",  f"{sub['accounts_per_device'].mean():.2f}", f"{normal['accounts_per_device'].mean():.2f}", f"{sub['accounts_per_device'].mean()/normal['accounts_per_device'].mean():.2f}x"),
            ("Avg sender_out_degree_30d",f"{sub['sender_out_degree_30d'].mean():.2f}", f"{normal['sender_out_degree_30d'].mean():.2f}", ""),
            ("Avg shared_device_fraud_count", f"{sub['shared_device_fraud_count'].mean():.2f}", f"{normal['shared_device_fraud_count'].mean():.2f}", ""),
            ("Top Channel",              sub['channel'].mode()[0],           normal['channel'].mode()[0],      ""),
            ("Top Transaction Type",     sub['transaction_type'].mode()[0],  normal['transaction_type'].mode()[0], ""),
        ]
        for metric, fval, nval, ratio in key_metrics:
            bg = C_LIGHT_BLUE if row%2==0 else C_WHITE
            for ci, v in enumerate([metric, fval, nval, ratio], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = body_font()
                c.fill = fill(bg)
                c.border = thin_border()
                c.alignment = left()
            row += 1

        # Top triggered rules
        top_rules = sub[rule_cols].mean().sort_values(ascending=False).head(5)
        ws.cell(row=row, column=1, value="Top Triggered Rules").font = body_font(bold=True)
        ws.cell(row=row, column=1).fill = fill(C_LIGHT_BLUE)
        ws.cell(row=row, column=1).border = thin_border()
        apply_header(ws, row, ["Top Triggered Rules","","Rate","","",""], bg=C_LIGHT_BLUE)
        row += 1
        for rule, rate in top_rules.items():
            for ci, v in enumerate([rule, "", f"{rate*100:.1f}%","","",""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = body_font()
                c.fill = fill(C_WHITE)
                c.border = thin_border()
                c.alignment = left()
            row += 1
        row += 2

    set_col_widths(ws, [(1,42),(2,22),(3,22),(4,14),(5,14),(6,14)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 10_Fraud_Type_Analysis")


# =============================================================================
# SHEET 11 – TOP CORRELATIONS
# =============================================================================
def sheet_correlations(wb, df):
    ws = wb.create_sheet("11_Top_Correlations")
    ws.freeze_panes = "A3"

    title_cell(ws, 1, 1, "Feature Correlations with Fraud Label")

    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    numeric_cols = [c for c in numeric_cols if c not in ["label","account_id_x","account_id_y"]]
    corr = df[numeric_cols + ["label"]].corr()["label"].drop("label").sort_values(key=abs, ascending=False)
    corr_df = corr.reset_index()
    corr_df.columns = ["feature","pearson_r"]
    corr_df["abs_r"] = corr_df["pearson_r"].abs()
    corr_df["direction"] = corr_df["pearson_r"].apply(lambda x: "Positive (↑ with fraud)" if x > 0 else "Negative (↓ with fraud)")

    # Feature group lookup
    corr_df["group"] = corr_df["feature"].map(lambda c: DATA_DICT.get(c, ("","","",""))[1])

    def signal_strength(r):
        if r > 0.5: return "Very Strong"
        if r > 0.3: return "Strong"
        if r > 0.15: return "Moderate"
        if r > 0.05: return "Weak"
        return "Negligible"

    corr_df["signal"] = corr_df["abs_r"].apply(signal_strength)

    apply_header(ws, 2, ["Rank","Feature","Feature Group","Pearson r","|r|","Direction","Signal Strength"], bg=C_DARK_BLUE)
    row = 3
    for rank, (_, r) in enumerate(corr_df.iterrows(), 1):
        abs_r = r["abs_r"]
        if abs_r > 0.5:   bg = "FCE4D6"
        elif abs_r > 0.3: bg = "FFF3E0"
        elif abs_r > 0.15:bg = "FFF9C4"
        else:             bg = C_WHITE
        vals = [rank, r["feature"], r["group"], round(r["pearson_r"],4), round(abs_r,4), r["direction"], r["signal"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = body_font(bold=(ci==2))
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        row += 1

    set_col_widths(ws, [(1,8),(2,42),(3,24),(4,14),(5,12),(6,28),(7,18)])
    ws.sheet_view.showGridLines = False
    print("  ✓ Sheet 11_Top_Correlations")


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(f"\nGenerating EDA Excel → {OUTPUT_PATH}\n")
    sheet_overview(wb, df)
    sheet_data_dict(wb, df)
    sheet_formulas(wb)
    sheet_desc_stats(wb, df)
    sheet_missing(wb, df)
    sheet_graph_eda(wb, df)
    sheet_categoricals(wb, df)
    sheet_rules(wb, df)
    sheet_temporal(wb, df)
    sheet_fraud_types(wb, df)
    sheet_correlations(wb, df)

    wb.save(OUTPUT_PATH)
    print(f"\n✅  Saved: {OUTPUT_PATH}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
